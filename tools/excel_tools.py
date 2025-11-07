"""
Excel Tools - XLSX parsing, extraction, and analysis tools.
Handles Excel spreadsheet processing for data analysis and reporting.
"""

import logging
import re
from pathlib import Path
from typing import Dict, Any, List, Optional

logger = logging.getLogger(__name__)

import pandas as pd
from openpyxl import load_workbook


def extract_sheets_from_xlsx(xlsx_path: str) -> List[str]:
    """
    Extract sheet names from Excel file.
    
    Args:
        xlsx_path: Path to XLSX file
        
    Returns:
        List of sheet names
    """
    logger.info(f"Extracting sheets from XLSX: {xlsx_path}")
    
    try:
        path = Path(xlsx_path)
        if not path.exists():
            return [f"Error: File not found: {xlsx_path}"]
        
        xl_file = pd.ExcelFile(xlsx_path)
        return xl_file.sheet_names
        
    except Exception as e:
        logger.error(f"Error extracting sheets from XLSX: {e}")
        return [f"Error: {str(e)}"]


def extract_data_from_xlsx(xlsx_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    Extract data from Excel file.
    
    Args:
        xlsx_path: Path to XLSX file
        sheet_name: Optional sheet name (default: first sheet)
        
    Returns:
        Dictionary with extracted data as list of dictionaries
    """
    logger.info(f"Extracting data from XLSX: {xlsx_path}, sheet: {sheet_name}")
    
    try:
        path = Path(xlsx_path)
        if not path.exists():
            return {"error": f"File not found: {xlsx_path}"}
        
        # Read the Excel file
        if sheet_name:
            df = pd.read_excel(xlsx_path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(xlsx_path)
        
        # Convert to dictionary format
        data_dict = {
            "sheet": sheet_name or "default",
            "columns": list(df.columns),
            "data": df.to_dict('records'),
            "row_count": len(df),
            "column_count": len(df.columns),
            "source": xlsx_path
        }
        
        return data_dict
        
    except Exception as e:
        logger.error(f"Error extracting data from XLSX: {e}")
        return {"error": str(e)}


def parse_batch_data_xlsx(xlsx_path: str) -> Dict[str, Any]:
    """
    Parse batch production data from Excel file.
    
    Args:
        xlsx_path: Path to batch data XLSX file
        
    Returns:
        Structured batch data with batch numbers, yields, dates
    """
    logger.info(f"Parsing batch data from XLSX: {xlsx_path}")
    
    try:
        # Get all sheet names
        sheet_names = extract_sheets_from_xlsx(xlsx_path)
        
        batch_data = {
            "document_type": "Batch Production Data",
            "source": xlsx_path,
            "sheets": sheet_names,
            "batches": [],
            "metadata": extract_metadata_from_xlsx(xlsx_path)
        }
        
        # Try to extract data from first sheet
        if sheet_names:
            first_sheet = sheet_names[0] if not sheet_names[0].startswith("Error") else None
            if first_sheet:
                df = pd.read_excel(xlsx_path, sheet_name=first_sheet)
                
                # Look for common batch data columns
                batch_columns = {}
                for col in df.columns:
                    col_lower = str(col).lower()
                    if 'batch' in col_lower or 'lot' in col_lower:
                        batch_columns['batch_number'] = col
                    elif 'yield' in col_lower:
                        batch_columns['yield'] = col
                    elif 'date' in col_lower:
                        batch_columns['date'] = col
                    elif 'product' in col_lower or 'material' in col_lower:
                        batch_columns['product'] = col
                    elif 'quantity' in col_lower or 'qty' in col_lower:
                        batch_columns['quantity'] = col
                
                # Extract batch records
                for idx, row in df.iterrows():
                    batch_record = {"row_number": idx + 2}  # +2 for header and 0-index
                    for key, col_name in batch_columns.items():
                        batch_record[key] = str(row[col_name]) if pd.notna(row[col_name]) else None
                    batch_data["batches"].append(batch_record)
        
        return batch_data
        
    except Exception as e:
        logger.error(f"Error parsing batch data XLSX: {e}")
        return {
            "document_type": "Batch Data",
            "error": str(e),
            "source": xlsx_path
        }


def parse_kpi_data_xlsx(xlsx_path: str) -> Dict[str, Any]:
    """
    Parse KPI data from Excel file.
    
    Args:
        xlsx_path: Path to KPI data XLSX file
        
    Returns:
        Structured KPI data with metrics, values, targets
    """
    logger.info(f"Parsing KPI data from XLSX: {xlsx_path}")
    
    try:
        # Get all sheet names
        sheet_names = extract_sheets_from_xlsx(xlsx_path)
        
        kpi_data = {
            "document_type": "KPI Dashboard",
            "source": xlsx_path,
            "sheets": sheet_names,
            "kpis": [],
            "metadata": extract_metadata_from_xlsx(xlsx_path)
        }
        
        # Try to extract KPI data from first sheet
        if sheet_names:
            first_sheet = sheet_names[0] if not sheet_names[0].startswith("Error") else None
            if first_sheet:
                df = pd.read_excel(xlsx_path, sheet_name=first_sheet)
                
                # Look for common KPI columns
                kpi_columns = {}
                for col in df.columns:
                    col_lower = str(col).lower()
                    if 'kpi' in col_lower or 'metric' in col_lower or 'indicator' in col_lower:
                        kpi_columns['kpi_name'] = col
                    elif 'target' in col_lower:
                        kpi_columns['target'] = col
                    elif 'actual' in col_lower or 'value' in col_lower:
                        kpi_columns['actual'] = col
                    elif 'status' in col_lower:
                        kpi_columns['status'] = col
                    elif 'unit' in col_lower or 'uom' in col_lower:
                        kpi_columns['unit'] = col
                
                # Extract KPI records
                for idx, row in df.iterrows():
                    kpi_record = {"row_number": idx + 2}
                    for key, col_name in kpi_columns.items():
                        kpi_record[key] = str(row[col_name]) if pd.notna(row[col_name]) else None
                    kpi_data["kpis"].append(kpi_record)
        
        return kpi_data
        
    except Exception as e:
        logger.error(f"Error parsing KPI data XLSX: {e}")
        return {
            "document_type": "KPI Dashboard",
            "error": str(e),
            "source": xlsx_path
        }


def extract_metadata_from_xlsx(xlsx_path: str) -> Dict[str, Any]:
    """
    Extract metadata from Excel file.
    
    Args:
        xlsx_path: Path to XLSX file
        
    Returns:
        Dictionary with XLSX metadata
    """
    logger.info(f"Extracting metadata from XLSX: {xlsx_path}")
    
    try:
        path = Path(xlsx_path)
        if path.exists():
            stat = path.stat()
            metadata = {
                "filename": path.name,
                "size_bytes": stat.st_size,
                "size_kb": round(stat.st_size / 1024, 2),
                "path": str(path),
                "exists": True
            }
            
            # Extract XLSX-specific metadata
            try:
                wb = load_workbook(xlsx_path, read_only=True)
                metadata["num_sheets"] = len(wb.sheetnames)
                metadata["sheet_names"] = wb.sheetnames
                
                # Try to get workbook properties
                if wb.properties:
                    metadata["xlsx_metadata"] = {
                        "creator": wb.properties.creator,
                        "created": str(wb.properties.created) if wb.properties.created else None,
                        "modified": str(wb.properties.modified) if wb.properties.modified else None,
                        "title": wb.properties.title,
                        "subject": wb.properties.subject
                    }
                wb.close()
            except Exception as e:
                logger.error(f"Error reading XLSX metadata: {e}")
            
            return metadata
        return {"filename": path.name, "exists": False}
    except Exception as e:
        logger.error(f"Error extracting metadata from XLSX: {e}")
        return {"error": str(e)}

